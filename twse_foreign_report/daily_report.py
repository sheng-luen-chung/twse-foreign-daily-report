#!/usr/bin/env python3
"""
外資買賣超日報工具 v4

功能升級：
1) --date 改為可省略；未提供時自動以台北時區今天為基準。
2) 自動回推最近交易日（預設最多回推 10 天），避開週末／休市日。
3) 保留每日歷史輸出，並維護 summary_history.csv。
4) 產出較漂亮的 Excel：自動欄寬、凍結窗格、數字格式、內嵌長條圖。
5) 另存「latest_*.csv / xlsx」，方便固定路徑給別的工具讀取。

主要資料源：
- TWSE TWT38U：外資及陸資買賣超彙總表
- TWSE MI_INDEX：每日收盤行情
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
import requests
import urllib3
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.exceptions import SSLError

from .shareholder_distribution import (
    build_shareholder_distribution_report,
    print_shareholder_report,
    save_shareholder_outputs,
    style_shareholder_sheets,
    write_shareholder_sheets,
)

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

FOREIGN_URLS = [
    "https://www.twse.com.tw/fund/TWT38U?date={date}&response=csv",
    "https://www.twse.com.tw/rwd/zh/fund/TWT38U?date={date}&response=csv",
]

QUOTE_JSON_URLS = [
    "https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date={date}&type=ALLBUT0999",
    "https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?response=json&date={date}&type=ALLBUT0999",
]

QUOTE_CSV_URLS = [
    "https://www.twse.com.tw/exchangeReport/MI_INDEX?response=csv&date={date}&type=ALLBUT0999",
    "https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?response=csv&date={date}&type=ALLBUT0999",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.twse.com.tw/",
}

CODE_PATTERN = re.compile(r"^[0-9A-Z]{4,6}$")
RECOVERABLE_HINTS = (
    "找不到外資買賣超表頭",
    "沒有解析出任何個股資料",
    "找不到每日收盤行情表",
    "CSV 中找不到『每日收盤行情』表頭",
    "JSON 中找不到每日收盤行情表",
    "網址回傳空白內容",
    "所有查詢網址都未回傳有效內容",
    "所有 JSON 查詢網址都未回傳有效內容",
    "抓取失敗，日期=",
    "JSON 抓取失敗，日期=",
    "no data",
    "沒有資料",
)


def normalize_security_code(value) -> str:
    """Normalize TWSE CSV codes such as ="00878 " back to 00878."""
    if value is None:
        return ""
    code = str(value).strip()
    if code.startswith("="):
        code = code[1:].strip()
    if len(code) >= 2 and code[0] == code[-1] == '"':
        code = code[1:-1].strip()
    return code.strip()


class NoDataError(RuntimeError):
    """該日期無有效交易資料。"""


@dataclass
class BuildResult:
    date: str
    merged: pd.DataFrame
    buys: pd.DataFrame
    sells: pd.DataFrame
    summary: pd.DataFrame


# -------------------------
# 基礎工具
# -------------------------

def today_in_taipei() -> str:
    if ZoneInfo is not None:
        now = datetime.now(ZoneInfo("Asia/Taipei"))
    else:  # pragma: no cover
        now = datetime.utcnow() + timedelta(hours=8)
    return now.strftime("%Y%m%d")



def normalize_date(date_str: Optional[str]) -> str:
    if not date_str:
        return today_in_taipei()
    s = re.sub(r"\D", "", str(date_str))
    if len(s) != 8:
        raise ValueError("日期格式請用 YYYYMMDD 或 YYYY-MM-DD")
    return s



def shift_date_yyyymmdd(date_str: str, days: int) -> str:
    dt = datetime.strptime(date_str, "%Y%m%d") + timedelta(days=days)
    return dt.strftime("%Y%m%d")



def to_number(x) -> float | None:
    if x is None:
        return None
    s = str(x).strip().replace(",", "")
    if s in {"", "--", "---", "----", "X", "除權", "除息", "暫停交易"}:
        return None
    s = s.replace("+", "")
    try:
        return float(s)
    except ValueError:
        return None



def signed_change(sign, delta) -> float | None:
    d = to_number(delta)
    if d is None:
        return None
    sign = str(sign).strip()
    if "-" in sign:
        return -abs(d)
    if "+" in sign:
        return abs(d)
    if str(delta).strip().startswith("-"):
        return -abs(d)
    return d



def is_recoverable_no_data(exc: Exception) -> bool:
    msg = str(exc)
    return any(hint in msg for hint in RECOVERABLE_HINTS)



def _get_with_ssl_fallback(url: str, timeout: int = 30):
    try:
        return requests.get(url, headers=HEADERS, timeout=timeout)
    except SSLError:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        return requests.get(url, headers=HEADERS, timeout=timeout, verify=False)



def fetch_text_with_fallback(url_templates: Iterable[str], date: str, timeout: int = 30) -> str:
    last_error: Optional[Exception] = None
    for tpl in url_templates:
        url = tpl.format(date=date)
        try:
            r = _get_with_ssl_fallback(url, timeout=timeout)
            r.raise_for_status()
            text = r.text.strip()
            if text:
                return text
            last_error = NoDataError(f"網址回傳空白內容：{url}")
        except Exception as e:
            last_error = e
        time.sleep(0.4)

    if last_error is None:
        last_error = NoDataError("所有查詢網址都未回傳有效內容，可能尚未開盤後更新或當日無資料。")
    raise NoDataError(f"抓取失敗，日期={date}，最後錯誤：{last_error}")



def fetch_json_with_fallback(url_templates: Iterable[str], date: str, timeout: int = 30) -> dict:
    last_error: Optional[Exception] = None
    for tpl in url_templates:
        url = tpl.format(date=date)
        try:
            r = _get_with_ssl_fallback(url, timeout=timeout)
            r.raise_for_status()
            data = r.json()
            if isinstance(data, dict) and data:
                stat = str(data.get("stat", "")).strip()
                if stat and ("沒有" in stat or "很抱歉" in stat or "查無" in stat or "No data" in stat):
                    last_error = NoDataError(f"{stat}（{url}）")
                else:
                    return data
            else:
                last_error = NoDataError(f"JSON 內容為空：{url}")
        except Exception as e:
            last_error = e
        time.sleep(0.4)

    if last_error is None:
        last_error = NoDataError("所有 JSON 查詢網址都未回傳有效內容，可能尚未開盤後更新或當日無資料。")
    raise NoDataError(f"JSON 抓取失敗，日期={date}，最後錯誤：{last_error}")



def find_exact_col_index(header: list[str], name: str, occurrence: int = 0) -> int:
    matches = [i for i, c in enumerate(header) if c.strip() == name]
    if len(matches) <= occurrence:
        raise KeyError(f"找不到欄位 {name!r} 第 {occurrence + 1} 次出現")
    return matches[occurrence]


# -------------------------
# 解析 TWSE 外資買賣超
# -------------------------

def parse_foreign_csv(text: str) -> pd.DataFrame:
    lines = [ln for ln in text.splitlines() if ln.strip() and ln.count(",") >= 5]
    rows = list(csv.reader(io.StringIO("\n".join(lines))))

    header_idx = None
    for i, row in enumerate(rows):
        joined = "|".join(row)
        if "證券代號" in joined and "證券名稱" in joined and "買賣超股數" in joined:
            header_idx = i
            break
    if header_idx is None:
        raise NoDataError("找不到外資買賣超表頭，可能是休市、查詢時間過早，或 TWSE 格式有變。")

    header = [c.strip() for c in rows[header_idx]]

    code_idx = find_exact_col_index(header, "證券代號")
    name_idx = find_exact_col_index(header, "證券名稱")
    buy_idx = find_exact_col_index(header, "買進股數", 0)
    sell_idx = find_exact_col_index(header, "賣出股數", 0)
    net_idx = find_exact_col_index(header, "買賣超股數", 0)

    records = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        row = row + [""] * (len(header) - len(row))
        code = normalize_security_code(row[code_idx])
        if not CODE_PATTERN.fullmatch(code):
            continue
        records.append(
            {
                "code": code,
                "name": row[name_idx].strip(),
                "foreign_buy_shares": to_number(row[buy_idx]),
                "foreign_sell_shares": to_number(row[sell_idx]),
                "foreign_net_shares": to_number(row[net_idx]),
            }
        )

    if not records:
        raise NoDataError("外資買賣超表有表頭，但沒有解析出任何個股資料。")

    out = pd.DataFrame.from_records(records)
    out["foreign_buy_lots"] = out["foreign_buy_shares"] / 1000.0
    out["foreign_sell_lots"] = out["foreign_sell_shares"] / 1000.0
    out["foreign_net_lots"] = out["foreign_net_shares"] / 1000.0
    return out


# -------------------------
# 解析 TWSE 收盤行情
# -------------------------

def clean_quotes_df(df: pd.DataFrame) -> pd.DataFrame:
    cols = {c.strip(): c for c in df.columns}
    code_col = next((cols[c] for c in cols if c == "證券代號"), None)
    name_col = next((cols[c] for c in cols if c == "證券名稱"), None)
    close_col = next((cols[c] for c in cols if c == "收盤價"), None)
    sign_col = next((cols[c] for c in cols if "漲跌(+/-)" in c), None)
    delta_col = next((cols[c] for c in cols if c in {"漲跌價差", "漲跌價差(元)"}), None)
    pct_col = next((cols[c] for c in cols if "漲跌幅" in c or "漲跌百分比" in c), None)

    if not (code_col and name_col and close_col):
        raise NoDataError("收盤行情表缺少必要欄位。")

    out = pd.DataFrame(
        {
            "code": df[code_col].map(normalize_security_code),
            "name_quote": df[name_col].astype(str).str.strip(),
            "close": df[close_col].map(to_number),
        }
    )

    if sign_col and delta_col:
        out["change"] = [signed_change(s, d) for s, d in zip(df[sign_col], df[delta_col])]
    elif delta_col:
        out["change"] = df[delta_col].map(to_number)
    else:
        out["change"] = None

    if pct_col:
        out["pct"] = df[pct_col].map(to_number)
    else:
        out["pct"] = None

    mask = out["pct"].isna() & out["close"].notna() & out["change"].notna()
    if mask.any():
        idx = out.index[mask]
        prev_close = out.loc[idx, "close"] - out.loc[idx, "change"]
        valid = prev_close != 0
        valid_idx = idx[valid]
        out.loc[valid_idx, "pct"] = out.loc[valid_idx, "change"] / prev_close[valid] * 100.0

    out = out[out["code"].str.fullmatch(r"[0-9A-Z]{4,6}", na=False)].copy()
    if out.empty:
        raise NoDataError("收盤行情表存在，但沒有解析出任何個股資料。")
    return out



def parse_quotes_json(data: dict) -> pd.DataFrame:
    tables = data.get("tables")
    if isinstance(tables, list):
        for tbl in tables:
            title = str(tbl.get("title", ""))
            fields = tbl.get("fields")
            rows = tbl.get("data")
            if not isinstance(fields, list) or not isinstance(rows, list) or not rows:
                continue
            if "證券代號" in fields and "證券名稱" in fields and "收盤價" in fields and "每日收盤行情" in title:
                return clean_quotes_df(pd.DataFrame(rows, columns=fields))

    for key, fields in data.items():
        if not key.startswith("fields") or not isinstance(fields, list):
            continue
        if "證券代號" not in fields or "證券名稱" not in fields or "收盤價" not in fields:
            continue
        data_key = "data" + key[len("fields") :]
        rows = data.get(data_key)
        if not isinstance(rows, list) or not rows:
            continue
        df = pd.DataFrame(rows, columns=fields)
        if len(df) > 100:
            return clean_quotes_df(df)

    raise NoDataError("JSON 中找不到每日收盤行情表，可能是休市、查詢時間過早，或格式改版。")



def parse_quotes_csv(text: str) -> pd.DataFrame:
    lines = [ln for ln in text.splitlines() if ln.strip()]
    header_idx = None
    for i, ln in enumerate(lines):
        if "證券代號" in ln and "證券名稱" in ln and "收盤價" in ln:
            header_idx = i
            break
    if header_idx is None:
        raise NoDataError("CSV 中找不到『每日收盤行情』表頭。")

    data_block = []
    for ln in lines[header_idx:]:
        if ln.startswith("=") or ("證券代號" not in ln and ln.count(",") < 5):
            if data_block:
                break
        data_block.append(ln)

    rows = list(csv.reader(io.StringIO("\n".join(data_block))))
    header = [c.strip() for c in rows[0]]
    body = []
    code_idx = find_exact_col_index(header, "證券代號")
    for row in rows[1:]:
        if not row:
            continue
        row = row + [""] * (len(header) - len(row))
        code = normalize_security_code(row[code_idx])
        if not CODE_PATTERN.fullmatch(code):
            continue
        body.append(row[: len(header)])

    if not body:
        raise NoDataError("CSV 每日收盤行情表存在，但沒有解析出任何個股資料。")

    df = pd.DataFrame(body, columns=header)
    return clean_quotes_df(df)



def fetch_quotes(date: str) -> pd.DataFrame:
    try:
        data = fetch_json_with_fallback(QUOTE_JSON_URLS, date)
        return parse_quotes_json(data)
    except Exception as exc:
        if not is_recoverable_no_data(exc):
            # JSON 結構錯誤、解析錯誤也可退回 CSV 再試一次
            pass
        text = fetch_text_with_fallback(QUOTE_CSV_URLS, date)
        return parse_quotes_csv(text)


# -------------------------
# 合併與統計
# -------------------------

def build_rank_table(date: str, top_n: int = 25) -> BuildResult:
    foreign_text = fetch_text_with_fallback(FOREIGN_URLS, date)
    foreign_df = parse_foreign_csv(foreign_text)
    quotes_df = fetch_quotes(date)

    merged = foreign_df.merge(quotes_df, on="code", how="left")
    if "name_quote" in merged.columns:
        merged["name"] = merged["name"].fillna(merged["name_quote"])
        merged.drop(columns=["name_quote"], inplace=True)

    merged["foreign_net_value_est"] = merged["foreign_net_lots"].fillna(0) * merged["close"].fillna(0) * 1000.0

    merged = merged.sort_values(["foreign_net_lots", "code"], ascending=[False, True]).reset_index(drop=True)

    buys = merged[merged["foreign_net_lots"] > 0].sort_values("foreign_net_lots", ascending=False).head(top_n).copy()
    sells = merged[merged["foreign_net_lots"] < 0].sort_values("foreign_net_lots", ascending=True).head(top_n).copy()

    buys.insert(0, "rank", range(1, len(buys) + 1))
    sells.insert(0, "rank", range(1, len(sells) + 1))

    summary = pd.DataFrame([
        {
            "date": date,
            "securities": int(len(merged)),
            "foreign_buy_lots_total": round(merged["foreign_buy_lots"].fillna(0).sum(), 2),
            "foreign_sell_lots_total": round(merged["foreign_sell_lots"].fillna(0).sum(), 2),
            "foreign_net_lots_total": round(merged["foreign_net_lots"].fillna(0).sum(), 2),
            "net_buy_count": int((merged["foreign_net_lots"] > 0).sum()),
            "net_sell_count": int((merged["foreign_net_lots"] < 0).sum()),
            "matched_close_count": int(merged["close"].notna().sum()),
            "net_buy_value_est_total": round(merged[merged["foreign_net_value_est"] > 0]["foreign_net_value_est"].sum(), 0),
            "net_sell_value_est_total": round(merged[merged["foreign_net_value_est"] < 0]["foreign_net_value_est"].sum(), 0),
            "top_buy_code": buys.iloc[0]["code"] if not buys.empty else None,
            "top_buy_name": buys.iloc[0]["name"] if not buys.empty else None,
            "top_buy_lots": round(float(buys.iloc[0]["foreign_net_lots"]), 2) if not buys.empty else None,
            "top_sell_code": sells.iloc[0]["code"] if not sells.empty else None,
            "top_sell_name": sells.iloc[0]["name"] if not sells.empty else None,
            "top_sell_lots": round(float(sells.iloc[0]["foreign_net_lots"]), 2) if not sells.empty else None,
        }
    ])

    return BuildResult(date=date, merged=merged, buys=buys, sells=sells, summary=summary)



def resolve_build(date: str, top_n: int, lookback_days: int, auto_prev: bool = True) -> BuildResult:
    attempted: list[str] = []
    current = date
    last_error: Optional[Exception] = None

    for _ in range(max(1, lookback_days + 1)):
        attempted.append(current)
        try:
            return build_rank_table(current, top_n)
        except Exception as exc:
            last_error = exc
            if not auto_prev:
                raise
            if is_recoverable_no_data(exc):
                current = shift_date_yyyymmdd(current, -1)
                continue
            raise

    attempt_msg = ", ".join(attempted)
    raise RuntimeError(f"近 {lookback_days} 天內都找不到可用資料。嘗試日期：{attempt_msg}；最後錯誤：{last_error}")


# -------------------------
# 輸出與 Excel 美化
# -------------------------

def select_report_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "rank",
        "code",
        "name",
        "foreign_buy_lots",
        "foreign_sell_lots",
        "foreign_net_lots",
        "close",
        "change",
        "pct",
        "foreign_net_value_est",
    ]
    existing = [c for c in cols if c in df.columns]
    return df[existing].copy()



def pretty_print(title: str, df: pd.DataFrame) -> None:
    view_cols = [c for c in ["rank", "code", "name", "foreign_net_lots", "close", "change", "pct"] if c in df.columns]
    shown = df[view_cols].copy()
    shown.columns = [
        {
            "rank": "名次",
            "code": "代號",
            "name": "名稱",
            "foreign_net_lots": "張數",
            "close": "收盤價",
            "change": "價差",
            "pct": "幅度(%)",
        }[c]
        for c in view_cols
    ]
    print(f"\n=== {title} ===")
    with pd.option_context(
        "display.max_rows", None,
        "display.max_columns", None,
        "display.width", 220,
        "display.unicode.east_asian_width", True,
    ):
        print(shown.to_string(index=False, justify="left"))



def auto_fit_worksheet(ws, max_width: int = 22) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value) + 2)
        ws.column_dimensions[letter].width = min(max(width, 10), max_width)



def style_worksheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    numeric_cols = {
        "D": "#,##0.00",
        "E": "#,##0.00",
        "F": "#,##0.00",
        "G": "#,##0.00",
        "H": "+0.00;-0.00;0.00",
        "I": "0.00%",
        "J": '#,##0',
    }
    for col, fmt in numeric_cols.items():
        for cell in ws[col][1:]:
            cell.number_format = fmt

    auto_fit_worksheet(ws)



def add_bar_chart(ws, title: str, data_col: int, label_col: int, anchor: str) -> None:
    if ws.max_row <= 1:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "股票"
    chart.x_axis.title = "張數"
    chart.height = 8
    chart.width = 14

    data = Reference(ws, min_col=data_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=label_col, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)



def build_dashboard_sheet(writer: pd.ExcelWriter, summary: pd.DataFrame, buys: pd.DataFrame, sells: pd.DataFrame) -> None:
    dashboard_rows = [
        ["日期", summary.iloc[0]["date"]],
        ["總檔數", summary.iloc[0]["securities"]],
        ["外資總買超張數", summary.iloc[0]["foreign_net_lots_total"]],
        ["買超檔數", summary.iloc[0]["net_buy_count"]],
        ["賣超檔數", summary.iloc[0]["net_sell_count"]],
        ["買超第一名", f"{summary.iloc[0]['top_buy_code']} {summary.iloc[0]['top_buy_name']} ({summary.iloc[0]['top_buy_lots']} 張)"],
        ["賣超第一名", f"{summary.iloc[0]['top_sell_code']} {summary.iloc[0]['top_sell_name']} ({summary.iloc[0]['top_sell_lots']} 張)"],
    ]
    dashboard_df = pd.DataFrame(dashboard_rows, columns=["項目", "值"])
    dashboard_df.to_excel(writer, sheet_name="dashboard", index=False)



def update_history(summary_history_path: Path, summary_row: pd.DataFrame) -> pd.DataFrame:
    incoming = summary_row.copy()
    if summary_history_path.exists():
        hist = pd.read_csv(summary_history_path, encoding="utf-8-sig")
        hist["date"] = hist["date"].astype(str)
        incoming["date"] = incoming["date"].astype(str)
        hist = pd.concat([hist, incoming], ignore_index=True)
        hist = hist.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    else:
        hist = incoming.copy()
    hist.to_csv(summary_history_path, index=False, encoding="utf-8-sig")
    return hist



def save_outputs(
    result: BuildResult,
    outdir: Path,
    keep_latest: bool = True,
    shareholder_result=None,
) -> dict[str, Path]:
    date = result.date
    outdir.mkdir(parents=True, exist_ok=True)
    archive_dir = outdir / "archive" / date
    history_dir = outdir / "history"
    archive_dir.mkdir(parents=True, exist_ok=True)
    history_dir.mkdir(parents=True, exist_ok=True)

    full_report = select_report_columns(result.merged)
    buy_report = select_report_columns(result.buys)
    sell_report = select_report_columns(result.sells)

    paths = {
        "full_csv": archive_dir / f"twse_foreign_full_{date}.csv",
        "top_buy_csv": archive_dir / f"twse_foreign_top_buy_{date}.csv",
        "top_sell_csv": archive_dir / f"twse_foreign_top_sell_{date}.csv",
        "summary_csv": archive_dir / f"twse_foreign_summary_{date}.csv",
        "xlsx": archive_dir / (
            f"twse_foreign_rank_with_holders_{date}.xlsx"
            if shareholder_result is not None
            else f"twse_foreign_rank_{date}.xlsx"
        ),
        "history_summary_csv": history_dir / "summary_history.csv",
    }

    full_report.to_csv(paths["full_csv"], index=False, encoding="utf-8-sig")
    buy_report.to_csv(paths["top_buy_csv"], index=False, encoding="utf-8-sig")
    sell_report.to_csv(paths["top_sell_csv"], index=False, encoding="utf-8-sig")
    result.summary.to_csv(paths["summary_csv"], index=False, encoding="utf-8-sig")

    hist = update_history(paths["history_summary_csv"], result.summary)

    shareholder_sheet_names: list[str] = []
    with pd.ExcelWriter(paths["xlsx"], engine="openpyxl") as writer:
        build_dashboard_sheet(writer, result.summary, buy_report, sell_report)
        result.summary.to_excel(writer, sheet_name="summary", index=False)
        buy_report.to_excel(writer, sheet_name="top_buy", index=False)
        sell_report.to_excel(writer, sheet_name="top_sell", index=False)
        full_report.to_excel(writer, sheet_name="full", index=False)
        hist.to_excel(writer, sheet_name="summary_history", index=False)
        if shareholder_result is not None:
            shareholder_sheet_names = write_shareholder_sheets(writer, shareholder_result, sheet_prefix="holders")

    wb = load_workbook(paths["xlsx"])
    for name in ["dashboard", "summary", "top_buy", "top_sell", "full", "summary_history"]:
        ws = wb[name]
        style_worksheet(ws)
    if shareholder_sheet_names:
        style_shareholder_sheets(wb, shareholder_sheet_names)

    # 百分比欄修正：目前 pct 是 3.5 代表 3.5%，Excel 百分比需 /100
    for sheet_name in ["top_buy", "top_sell", "full"]:
        ws = wb[sheet_name]
        headers = {cell.value: cell.column for cell in ws[1]}
        if "pct" in headers:
            col = headers["pct"]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100.0
                cell.number_format = "0.00%"

    if "top_buy" in wb.sheetnames:
        ws_buy = wb["top_buy"]
        headers = {cell.value: cell.column for cell in ws_buy[1]}
        if "foreign_net_lots" in headers and "name" in headers:
            add_bar_chart(ws_buy, "外資買超前段排行榜", headers["foreign_net_lots"], headers["name"], "L2")

    if "top_sell" in wb.sheetnames:
        ws_sell = wb["top_sell"]
        headers = {cell.value: cell.column for cell in ws_sell[1]}
        if "foreign_net_lots" in headers and "name" in headers:
            add_bar_chart(ws_sell, "外資賣超前段排行榜", headers["foreign_net_lots"], headers["name"], "L2")

    wb.save(paths["xlsx"])

    if keep_latest:
        latest_map = {
            "latest_full.csv": paths["full_csv"],
            "latest_top_buy.csv": paths["top_buy_csv"],
            "latest_top_sell.csv": paths["top_sell_csv"],
            "latest_summary.csv": paths["summary_csv"],
            "latest_report.xlsx": paths["xlsx"],
        }
        for latest_name, src in latest_map.items():
            shutil.copy2(src, outdir / latest_name)

    return paths


# -------------------------
# CLI
# -------------------------

def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="TWSE 外資日報 / TDCC 股權分散工具")
    p.add_argument("--date", help="日期，格式 YYYYMMDD 或 YYYY-MM-DD；省略則以台北今天為基準")
    p.add_argument("--top", type=int, default=25, help="排行榜筆數，預設 25")
    p.add_argument("--outdir", default="output", help="輸出資料夾，預設 output")
    p.add_argument("--lookback", type=int, default=10, help="自動往前回推最近交易日的最大天數，預設 10")
    p.add_argument("--no-auto-prev", action="store_true", help="不要自動往前找最近交易日")
    p.add_argument(
        "--holders-targets",
        nargs="+",
        help="查詢集保戶股權分散表，可輸入股票代號或名稱，例如 --holders-targets 鴻海 臻鼎",
    )
    p.add_argument("--holders-weeks", type=int, default=3, help="股權分散表比較週數，預設 3")
    return p.parse_args(argv)



def main(argv: list[str]) -> int:
    args = parse_args(argv)
    base_date = normalize_date(args.date)
    outdir = Path(args.outdir)

    if args.holders_targets:
        foreign_result = resolve_build(
            date=base_date,
            top_n=args.top,
            lookback_days=max(0, args.lookback),
            auto_prev=not args.no_auto_prev,
        )
        shareholder_result = build_shareholder_distribution_report(
            base_date=base_date,
            targets=args.holders_targets,
            weeks=args.holders_weeks,
        )
        holder_paths = save_shareholder_outputs(shareholder_result, outdir, keep_latest=True)
        foreign_paths = save_outputs(
            foreign_result,
            outdir,
            keep_latest=True,
            shareholder_result=shareholder_result,
        )

        print(f"查詢基準日：{base_date}")
        print(f"外資實際使用交易日：{foreign_result.date}")
        print(f"股權分散表週資料：{' / '.join(shareholder_result.selected_dates)}")
        print("查詢股票：" + ", ".join(f"{item.code} {item.name}" for item in shareholder_result.targets))
        print_shareholder_report(shareholder_result, include_header=False)
        print("\n主要輸出：")
        print(f"- 整合 Excel：{foreign_paths['xlsx'].resolve()}")
        print(f"- 股權分散表 Excel：{holder_paths['xlsx'].resolve()}")
        print(f"- 明細 CSV：{holder_paths['detail_csv'].resolve()}")
        print(f"- 最近兩週變化 CSV：{holder_paths['recent_changes_csv'].resolve()}")
        print(f"- 變化 CSV：{holder_paths['changes_csv'].resolve()}")
        print(f"- latest_report.xlsx：{(outdir / 'latest_report.xlsx').resolve()}")
        print(f"- latest_shareholders_report.xlsx：{(outdir / 'latest_shareholders_report.xlsx').resolve()}")
        return 0

    result = resolve_build(
        date=base_date,
        top_n=args.top,
        lookback_days=max(0, args.lookback),
        auto_prev=not args.no_auto_prev,
    )
    paths = save_outputs(result, outdir, keep_latest=True)

    print(f"查詢基準日：{base_date}")
    print(f"實際使用交易日：{result.date}")
    print(result.summary.to_string(index=False))
    pretty_print(f"外資買超前 {len(result.buys)} 名", result.buys)
    pretty_print(f"外資賣超前 {len(result.sells)} 名", result.sells)
    print("\n主要輸出：")
    print(f"- Excel：{paths['xlsx'].resolve()}")
    print(f"- 歷史摘要：{paths['history_summary_csv'].resolve()}")
    print(f"- latest_report.xlsx：{(outdir / 'latest_report.xlsx').resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
