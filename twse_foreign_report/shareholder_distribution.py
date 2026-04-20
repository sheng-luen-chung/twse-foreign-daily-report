from __future__ import annotations

import html
import re
import shutil
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.exceptions import SSLError

TDCC_QUERY_URL = "https://www.tdcc.com.tw/portal/zh/smWeb/qryStock"
ISIN_SEARCH_URL = "https://isin.twse.com.tw/isin/class_main.jsp"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.tdcc.com.tw/",
}

CODE_PATTERN = re.compile(r"^[0-9A-Z]{4,6}$")
ROW_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.IGNORECASE | re.DOTALL)
CELL_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.IGNORECASE | re.DOTALL)
TAG_RE = re.compile(r"<[^>]+>")

HOLDING_LEVEL_LABELS = {
    1: "1-999",
    2: "1,000-5,000",
    3: "5,001-10,000",
    4: "10,001-15,000",
    5: "15,001-20,000",
    6: "20,001-30,000",
    7: "30,001-40,000",
    8: "40,001-50,000",
    9: "50,001-100,000",
    10: "100,001-200,000",
    11: "200,001-400,000",
    12: "400,001-600,000",
    13: "600,001-800,000",
    14: "800,001-1,000,000",
    15: "1,000,001以上",
    16: "合計",
}

DETAIL_LEVELS = list(range(1, 16))
TOTAL_LEVEL = 16


class ShareholderDataError(RuntimeError):
    """Raised when TDCC shareholder distribution data is unavailable or invalid."""


@dataclass(frozen=True)
class ResolvedTarget:
    query: str
    code: str
    name: str


@dataclass
class ShareholderBuildResult:
    base_date: str
    selected_dates: list[str]
    targets: list[ResolvedTarget]
    detail: pd.DataFrame
    changes: pd.DataFrame
    recent_changes: pd.DataFrame
    summary: pd.DataFrame


def _normalize_date(value: str) -> str:
    s = re.sub(r"\D", "", str(value or ""))
    if len(s) != 8:
        raise ValueError(f"日期格式錯誤：{value!r}")
    return s


def _normalize_key(value: str) -> str:
    return (
        str(value or "")
        .strip()
        .upper()
        .replace("－", "-")
        .replace("–", "-")
        .replace("—", "-")
        .replace(" ", "")
        .replace("\u3000", "")
    )


QUERY_RANGE_TO_LEVEL = {
    _normalize_key(label): level for level, label in HOLDING_LEVEL_LABELS.items()
}


def _to_int(value: str) -> int:
    s = str(value or "").strip().replace(",", "")
    if not s:
        return 0
    return int(float(s))


def _to_float(value: str) -> float:
    s = str(value or "").strip().replace(",", "")
    if not s:
        return 0.0
    return float(s)


def _get_with_ssl_fallback(url: str, **kwargs):
    try:
        return requests.get(url, headers=HEADERS, **kwargs)
    except SSLError:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        return requests.get(url, headers=HEADERS, verify=False, **kwargs)


def _strip_html(value: str) -> str:
    text = re.sub(r"<br\s*/?>", " ", value, flags=re.IGNORECASE)
    text = TAG_RE.sub("", text)
    return " ".join(html.unescape(text).split())


def _extract_isin_rows(html_text: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_html in ROW_RE.findall(html_text):
        cells = [_strip_html(cell) for cell in CELL_RE.findall(row_html)]
        if len(cells) != 10:
            continue
        if not cells[0].isdigit():
            continue
        rows.append(
            {
                "page_no": cells[0],
                "isin_code": cells[1],
                "code": cells[2],
                "name": cells[3],
                "market": cells[4],
                "security_type": cells[5],
                "industry": cells[6],
                "listed_date": cells[7],
                "cfi_code": cells[8],
                "remark": cells[9],
            }
        )
    return rows


def _prefer_equity_rows(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    preferred: list[dict[str, str]] = []
    for row in rows:
        market = row["market"].strip()
        security_type = row["security_type"].strip()
        if market not in {"上市", "上櫃", "興櫃"}:
            continue
        if "股票" in security_type or "存託憑證" in security_type:
            preferred.append(row)
    return preferred or list(rows)


@lru_cache(maxsize=128)
def _search_isin(query: str, by_name: bool) -> tuple[tuple[str, str, str, str], ...]:
    params = {
        "owncode": "" if by_name else query,
        "stockname": query if by_name else "",
        "isincode": "",
        "market": "",
        "issuetype": "",
        "industry_code": "",
        "Page": "1",
        "chklike": "Y",
    }
    response = _get_with_ssl_fallback(ISIN_SEARCH_URL, params=params, timeout=30)
    response.raise_for_status()
    rows = _prefer_equity_rows(_extract_isin_rows(response.text))
    return tuple((row["code"], row["name"], row["market"], row["security_type"]) for row in rows)


def _resolve_target(query: str) -> ResolvedTarget:
    normalized = _normalize_key(query)
    if not normalized:
        raise ValueError("股票代號或名稱不能空白。")

    by_name = not CODE_PATTERN.fullmatch(normalized)
    candidates = _search_isin(normalized if by_name else normalized, by_name=by_name)
    if not candidates:
        raise ValueError(f"找不到股票：{query}")

    if not by_name:
        exact = [row for row in candidates if row[0] == normalized]
        chosen = exact[0] if exact else candidates[0]
        return ResolvedTarget(query=query, code=chosen[0], name=chosen[1])

    exact_name = [row for row in candidates if _normalize_key(row[1]) == normalized]
    if exact_name:
        chosen = exact_name[0]
        return ResolvedTarget(query=query, code=chosen[0], name=chosen[1])

    contains = [row for row in candidates if normalized in _normalize_key(row[1])]
    if len(contains) == 1:
        chosen = contains[0]
        return ResolvedTarget(query=query, code=chosen[0], name=chosen[1])

    unique_codes = {row[0]: row for row in contains or candidates}
    if len(unique_codes) == 1:
        chosen = next(iter(unique_codes.values()))
        return ResolvedTarget(query=query, code=chosen[0], name=chosen[1])

    sample = ", ".join(f"{code} {name}" for code, name, _, _ in list(unique_codes.values())[:5])
    raise ValueError(f"股票名稱 {query!r} 對到多筆結果，請改用代號或更完整名稱，例如：{sample}")


def resolve_targets(targets: list[str]) -> list[ResolvedTarget]:
    if not targets:
        raise ValueError("請至少指定一檔股票，例如 --holders-targets 鴻海 臻鼎")

    resolved: list[ResolvedTarget] = []
    seen_codes: set[str] = set()
    for target in targets:
        item = _resolve_target(target)
        if item.code in seen_codes:
            continue
        resolved.append(item)
        seen_codes.add(item.code)
    return resolved


def _extract_tdcc_context(html_text: str) -> dict[str, object]:
    token_match = re.search(r'name="SYNCHRONIZER_TOKEN" value="([^"]+)"', html_text)
    uri_match = re.search(r'name="SYNCHRONIZER_URI" value="([^"]+)"', html_text)
    fir_date_match = re.search(r'name="firDate" value="([^"]+)"', html_text)
    available_dates = re.findall(r'<option value="(\d{8})"', html_text)

    if not (token_match and uri_match and fir_date_match):
        raise ShareholderDataError("TDCC 查詢頁格式有變，找不到查詢表單欄位。")
    if not available_dates:
        raise ShareholderDataError("TDCC 查詢頁格式有變，找不到可用資料日期。")

    return {
        "token": token_match.group(1),
        "uri": uri_match.group(1),
        "fir_date": fir_date_match.group(1),
        "available_dates": list(dict.fromkeys(available_dates)),
    }


def _extract_query_rows(html_text: str, code: str, date: str) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for row_html in ROW_RE.findall(html_text):
        cells = [_strip_html(cell) for cell in CELL_RE.findall(row_html)]
        if len(cells) != 5:
            continue
        if not cells[0].isdigit():
            continue

        holding_level = QUERY_RANGE_TO_LEVEL.get(_normalize_key(cells[1]))
        if holding_level is None:
            continue

        records.append(
            {
                "date": date,
                "code": code,
                "holding_level": holding_level,
                "holders": _to_int(cells[2]),
                "shares": _to_int(cells[3]),
                "ratio_pct": _to_float(cells[4]),
            }
        )
    return records


def _fetch_filtered_shareholder_rows(
    codes: set[str],
    base_date: str,
    weeks: int,
) -> tuple[pd.DataFrame, list[str]]:
    session = requests.Session()
    response = session.get(TDCC_QUERY_URL, headers=HEADERS, timeout=30)
    response.raise_for_status()
    context = _extract_tdcc_context(response.text)

    available_dates = [date for date in context["available_dates"] if date <= base_date]
    selected_dates = available_dates[:weeks]
    if not selected_dates:
        raise ShareholderDataError(f"TDCC 找不到 {base_date} 當週或更早的股權分散表資料。")

    records: list[dict[str, object]] = []

    for code in sorted(codes):
        for date in selected_dates:
            payload = {
                "SYNCHRONIZER_TOKEN": context["token"],
                "SYNCHRONIZER_URI": context["uri"],
                "method": "submit",
                "firDate": context["fir_date"],
                "scaDate": date,
                "sqlMethod": "StockNo",
                "stockNo": code,
                "stockName": "",
            }
            result = session.post(TDCC_QUERY_URL, data=payload, headers=HEADERS, timeout=30)
            result.raise_for_status()
            context = _extract_tdcc_context(result.text)
            rows = _extract_query_rows(result.text, code=code, date=date)
            if not rows:
                raise ShareholderDataError(f"TDCC 查不到 {code} 在 {date} 的股權分散表資料。")
            records.extend(rows)

    if not records:
        raise ShareholderDataError("指定股票在最近可用週資料中查不到股權分散表。")

    return pd.DataFrame.from_records(records), selected_dates


def _complete_detail_grid(detail: pd.DataFrame, targets: list[ResolvedTarget], selected_dates: list[str]) -> pd.DataFrame:
    full_index = pd.MultiIndex.from_product(
        [[target.code for target in targets], selected_dates, DETAIL_LEVELS],
        names=["code", "date", "holding_level"],
    )
    completed = detail.set_index(["code", "date", "holding_level"]).reindex(full_index).reset_index()
    name_map = {target.code: target.name for target in targets}
    completed["name"] = completed["code"].map(name_map)
    completed["holders"] = completed["holders"].fillna(0).astype(int)
    completed["shares"] = completed["shares"].fillna(0).astype(int)
    completed["ratio_pct"] = completed["ratio_pct"].fillna(0.0).astype(float)
    completed["holding_range"] = completed["holding_level"].map(HOLDING_LEVEL_LABELS)
    completed = completed.sort_values(["code", "date", "holding_level"]).reset_index(drop=True)
    return completed


def _build_change_table(detail: pd.DataFrame, selected_dates: list[str]) -> pd.DataFrame:
    index_cols = ["code", "name", "holding_level", "holding_range"]
    base = detail[index_cols].drop_duplicates().sort_values(["code", "holding_level"]).reset_index(drop=True)

    for metric in ["holders", "shares", "ratio_pct"]:
        pivot = detail.pivot_table(index=index_cols, columns="date", values=metric, aggfunc="first")
        pivot = pivot.reindex(columns=selected_dates)
        pivot.columns = [f"{metric}_{date}" for date in selected_dates]
        base = base.merge(pivot.reset_index(), on=index_cols, how="left")

    for newer, older in zip(selected_dates, selected_dates[1:]):
        base[f"holders_change_{newer}_vs_{older}"] = base[f"holders_{newer}"] - base[f"holders_{older}"]
        base[f"shares_change_{newer}_vs_{older}"] = base[f"shares_{newer}"] - base[f"shares_{older}"]
        base[f"lots_change_{newer}_vs_{older}"] = base[f"shares_change_{newer}_vs_{older}"] / 1000.0

    int_prefixes = ("holders_", "shares_")
    for col in base.columns:
        if col.startswith(int_prefixes) and "ratio_pct_" not in col:
            base[col] = base[col].fillna(0).astype(int)
        elif col.startswith("ratio_pct_"):
            base[col] = base[col].fillna(0.0).astype(float)
        elif col.startswith("lots_change_"):
            base[col] = base[col].fillna(0.0).astype(float)

    for date in selected_dates:
        base[f"shares_lots_{date}"] = base[f"shares_{date}"] / 1000.0

    for col in base.columns:
        if col.startswith("ratio_pct_"):
            base[col] = base[col].round(2)
        elif col.startswith("shares_lots_") or col.startswith("lots_change_"):
            base[col] = base[col].round(3)

    return base


def _build_recent_change_table(detail: pd.DataFrame, selected_dates: list[str]) -> pd.DataFrame:
    if len(selected_dates) < 2:
        raise ValueError("至少需要兩週資料才能建立最近兩週變化表。")

    latest, previous = selected_dates[0], selected_dates[1]
    index_cols = ["code", "name", "holding_level", "holding_range"]
    base = detail[index_cols].drop_duplicates().sort_values(["code", "holding_level"]).reset_index(drop=True)

    for metric in ["holders", "shares", "ratio_pct"]:
        pivot = detail.pivot_table(index=index_cols, columns="date", values=metric, aggfunc="first")
        pivot = pivot.reindex(columns=[latest, previous])
        base = base.merge(pivot.reset_index(), on=index_cols, how="left")
        base.rename(
            columns={
                latest: f"{metric}_{latest}",
                previous: f"{metric}_{previous}",
            },
            inplace=True,
        )

    base[f"holders_change_{latest}_vs_{previous}"] = base[f"holders_{latest}"] - base[f"holders_{previous}"]
    base[f"shares_change_{latest}_vs_{previous}"] = base[f"shares_{latest}"] - base[f"shares_{previous}"]
    base[f"shares_lots_{latest}"] = base[f"shares_{latest}"] / 1000.0
    base[f"shares_lots_{previous}"] = base[f"shares_{previous}"] / 1000.0
    base[f"lots_change_{latest}_vs_{previous}"] = base[f"shares_change_{latest}_vs_{previous}"] / 1000.0
    base[f"ratio_pct_change_{latest}_vs_{previous}"] = base[f"ratio_pct_{latest}"] - base[f"ratio_pct_{previous}"]

    int_cols = [
        f"holders_{latest}",
        f"holders_{previous}",
        f"shares_{latest}",
        f"shares_{previous}",
        f"holders_change_{latest}_vs_{previous}",
        f"shares_change_{latest}_vs_{previous}",
    ]
    float_cols = [
        f"shares_lots_{latest}",
        f"shares_lots_{previous}",
        f"lots_change_{latest}_vs_{previous}",
        f"ratio_pct_{latest}",
        f"ratio_pct_{previous}",
        f"ratio_pct_change_{latest}_vs_{previous}",
    ]
    for col in int_cols:
        base[col] = base[col].fillna(0).astype(int)
    for col in float_cols:
        base[col] = base[col].fillna(0.0).astype(float)

    for col in [f"shares_lots_{latest}", f"shares_lots_{previous}", f"lots_change_{latest}_vs_{previous}"]:
        base[col] = base[col].round(3)
    for col in [f"ratio_pct_{latest}", f"ratio_pct_{previous}", f"ratio_pct_change_{latest}_vs_{previous}"]:
        base[col] = base[col].round(2)

    return base


def _build_summary_table(raw: pd.DataFrame, targets: list[ResolvedTarget], selected_dates: list[str]) -> pd.DataFrame:
    totals = raw[raw["holding_level"] == TOTAL_LEVEL].copy()
    if totals.empty:
        totals = pd.DataFrame(columns=["date", "code", "holding_level", "holders", "shares", "ratio_pct"])

    full_index = pd.MultiIndex.from_product(
        [[target.code for target in targets], selected_dates],
        names=["code", "date"],
    )
    totals = totals.set_index(["code", "date"]).reindex(full_index).reset_index()
    name_map = {target.code: target.name for target in targets}
    totals["name"] = totals["code"].map(name_map)
    totals["holders"] = totals["holders"].fillna(0).astype(int)
    totals["shares"] = totals["shares"].fillna(0).astype(int)

    summary = totals[["code", "name"]].drop_duplicates().sort_values("code").reset_index(drop=True)
    for metric in ["holders", "shares"]:
        pivot = totals.pivot_table(index=["code", "name"], columns="date", values=metric, aggfunc="first")
        pivot = pivot.reindex(columns=selected_dates)
        prefix = "total_holders" if metric == "holders" else "total_shares"
        pivot.columns = [f"{prefix}_{date}" for date in selected_dates]
        summary = summary.merge(pivot.reset_index(), on=["code", "name"], how="left")

    for newer, older in zip(selected_dates, selected_dates[1:]):
        summary[f"total_holders_change_{newer}_vs_{older}"] = (
            summary[f"total_holders_{newer}"] - summary[f"total_holders_{older}"]
        )
        summary[f"total_shares_change_{newer}_vs_{older}"] = (
            summary[f"total_shares_{newer}"] - summary[f"total_shares_{older}"]
        )
        summary[f"total_lots_change_{newer}_vs_{older}"] = (
            summary[f"total_shares_change_{newer}_vs_{older}"] / 1000.0
        )

    for col in summary.columns:
        if col.startswith("total_"):
            if "_lots_" in col:
                summary[col] = summary[col].fillna(0.0).astype(float)
            else:
                summary[col] = summary[col].fillna(0).astype(int)

    for date in selected_dates:
        summary[f"total_lots_{date}"] = summary[f"total_shares_{date}"] / 1000.0

    for col in summary.columns:
        if col.startswith("total_lots_"):
            summary[col] = summary[col].round(3)

    return summary


def build_shareholder_distribution_report(
    base_date: str,
    targets: list[str],
    weeks: int = 3,
) -> ShareholderBuildResult:
    if weeks < 2:
        raise ValueError("--holders-weeks 至少要 2，這樣才看得到變化。")

    resolved_targets = resolve_targets(targets)
    raw, selected_dates = _fetch_filtered_shareholder_rows(
        codes={target.code for target in resolved_targets},
        base_date=_normalize_date(base_date),
        weeks=weeks,
    )

    name_map = {target.code: target.name for target in resolved_targets}
    raw["name"] = raw["code"].map(name_map)
    raw["holding_range"] = raw["holding_level"].map(HOLDING_LEVEL_LABELS)

    detail_raw = raw[raw["holding_level"].isin(DETAIL_LEVELS)].copy()
    missing_codes = sorted(set(name_map) - set(detail_raw["code"].unique()))
    if missing_codes:
        names = ", ".join(f"{code} {name_map[code]}" for code in missing_codes)
        raise ShareholderDataError(f"最近 {len(selected_dates)} 週資料缺少以下股票：{names}")

    detail = _complete_detail_grid(
        detail_raw,
        targets=resolved_targets,
        selected_dates=selected_dates,
    )
    changes = _build_change_table(detail, selected_dates)
    recent_changes = _build_recent_change_table(detail, selected_dates)
    summary = _build_summary_table(raw, resolved_targets, selected_dates)

    return ShareholderBuildResult(
        base_date=_normalize_date(base_date),
        selected_dates=selected_dates,
        targets=resolved_targets,
        detail=detail,
        changes=changes,
        recent_changes=recent_changes,
        summary=summary,
    )


def auto_fit_worksheet(ws, max_width: int = 24) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value) + 2)
        ws.column_dimensions[letter].width = min(max(width, 10), max_width)


def style_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '#,##0'
            elif isinstance(cell.value, float):
                cell.number_format = "0.00"

    auto_fit_worksheet(ws)


def build_shareholder_overview(result: ShareholderBuildResult) -> pd.DataFrame:
    overview_rows = [
        ["查詢基準日", result.base_date],
        ["實際週資料", " / ".join(result.selected_dates)],
        ["股票", ", ".join(f"{item.code} {item.name}" for item in result.targets)],
        ["資料來源", "TDCC 集保戶股權分散表查詢頁"],
    ]
    return pd.DataFrame(overview_rows, columns=["項目", "值"])


def write_shareholder_sheets(
    writer: pd.ExcelWriter,
    result: ShareholderBuildResult,
    sheet_prefix: str = "",
) -> list[str]:
    prefix = f"{sheet_prefix}_" if sheet_prefix else ""
    sheets = {
        f"{prefix}overview": build_shareholder_overview(result),
        f"{prefix}summary": result.summary,
        f"{prefix}recent_changes": result.recent_changes,
        f"{prefix}changes": result.changes,
        f"{prefix}detail": result.detail,
    }
    for sheet_name, df in sheets.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return list(sheets.keys())


def style_shareholder_sheets(workbook, sheet_names: Iterable[str]) -> None:
    for sheet_name in sheet_names:
        style_worksheet(workbook[sheet_name])


def save_shareholder_outputs(result: ShareholderBuildResult, outdir: Path, keep_latest: bool = True) -> dict[str, Path]:
    outdir.mkdir(parents=True, exist_ok=True)
    latest_date = result.selected_dates[0]
    archive_dir = outdir / "archive" / latest_date
    archive_dir.mkdir(parents=True, exist_ok=True)
    week_suffix = f"w{len(result.selected_dates)}"

    paths = {
        "detail_csv": archive_dir / f"tdcc_shareholders_detail_{latest_date}_{week_suffix}.csv",
        "recent_changes_csv": archive_dir / f"tdcc_shareholders_recent_changes_{latest_date}_{week_suffix}.csv",
        "changes_csv": archive_dir / f"tdcc_shareholders_changes_{latest_date}_{week_suffix}.csv",
        "summary_csv": archive_dir / f"tdcc_shareholders_summary_{latest_date}_{week_suffix}.csv",
        "xlsx": archive_dir / f"tdcc_shareholders_report_{latest_date}_{week_suffix}.xlsx",
    }

    result.detail.to_csv(paths["detail_csv"], index=False, encoding="utf-8-sig")
    result.recent_changes.to_csv(paths["recent_changes_csv"], index=False, encoding="utf-8-sig")
    result.changes.to_csv(paths["changes_csv"], index=False, encoding="utf-8-sig")
    result.summary.to_csv(paths["summary_csv"], index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(paths["xlsx"], engine="openpyxl") as writer:
        sheet_names = write_shareholder_sheets(writer, result)

    workbook = load_workbook(paths["xlsx"])
    style_shareholder_sheets(workbook, sheet_names)
    workbook.save(paths["xlsx"])

    if keep_latest:
        latest_map = {
            "latest_shareholders_detail.csv": paths["detail_csv"],
            "latest_shareholders_recent_changes.csv": paths["recent_changes_csv"],
            "latest_shareholders_changes.csv": paths["changes_csv"],
            "latest_shareholders_summary.csv": paths["summary_csv"],
            "latest_shareholders_report.xlsx": paths["xlsx"],
        }
        for latest_name, source in latest_map.items():
            shutil.copy2(source, outdir / latest_name)

    return paths


def print_shareholder_report(result: ShareholderBuildResult, include_header: bool = True) -> None:
    if include_header:
        print(f"查詢基準日：{result.base_date}")
        print(f"實際使用週資料：{' / '.join(result.selected_dates)}")
        print("查詢股票：" + ", ".join(f"{item.code} {item.name}" for item in result.targets))

    latest, previous = result.selected_dates[0], result.selected_dates[1]
    preview_cols = [
        "holding_range",
        f"holders_{latest}",
        f"holders_{previous}",
        f"holders_change_{latest}_vs_{previous}",
        f"shares_lots_{latest}",
        f"shares_lots_{previous}",
        f"lots_change_{latest}_vs_{previous}",
    ]

    for item in result.targets:
        print(f"\n=== {item.code} {item.name} ===")
        target_df = result.recent_changes[result.recent_changes["code"] == item.code][preview_cols].copy()
        rename_map = {"holding_range": "持股分級"}
        rename_map.update(
            {
                f"holders_{latest}": f"{latest} 人數",
                f"holders_{previous}": f"{previous} 人數",
                f"holders_change_{latest}_vs_{previous}": f"人數增減({latest} vs {previous})",
                f"shares_lots_{latest}": f"{latest} 張數",
                f"shares_lots_{previous}": f"{previous} 張數",
                f"lots_change_{latest}_vs_{previous}": f"張數增減({latest} vs {previous})",
            }
        )
        target_df.rename(columns=rename_map, inplace=True)
        with pd.option_context(
            "display.max_rows", None,
            "display.max_columns", None,
            "display.width", 220,
            "display.unicode.east_asian_width", True,
        ):
            print(target_df.to_string(index=False))
